import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy as copy_obj
import io
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(page_title="Excel Transpose Tool", page_icon="🔄", layout="centered")

st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;600&display=swap');
        html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
        .block-container { padding-top: 2rem; max-width: 780px; }
        h1 { font-family: 'Space Mono', monospace; font-size: 1.8rem !important; letter-spacing: -1px; color: #0f172a; }
        .subtitle { color: #64748b; font-size: 0.95rem; margin-top: -0.5rem; margin-bottom: 2rem; }
        .step-card { background: #f8fafc; border: 1px solid #e2e8f0; border-left: 4px solid #6366f1; border-radius: 8px; padding: 1rem 1.2rem; margin-bottom: 0.8rem; font-size: 0.9rem; color: #334155; }
        .step-card b { color: #6366f1; font-family: 'Space Mono', monospace; }
        .success-box { background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 8px; padding: 1rem 1.2rem; color: #166534; font-size: 0.9rem; }
        .stDownloadButton > button { background-color: #6366f1 !important; color: white !important; border: none !important; padding: 0.6rem 1.4rem !important; border-radius: 6px !important; font-family: 'Space Mono', monospace !important; font-size: 0.85rem !important; width: 100%; }
        .stDownloadButton > button:hover { background-color: #4f46e5 !important; }
        footer { visibility: hidden; }
    </style>
""", unsafe_allow_html=True)

st.markdown("<h1>🔄 Excel Transpose Tool</h1>", unsafe_allow_html=True)
st.markdown('<p class="subtitle">Upload your Excel file to transpose it while preserving all colors and styles.</p>', unsafe_allow_html=True)

st.markdown("""
<div class="step-card"><b>Step 1</b> &nbsp;Unmerge all cells & fill missing values</div>
<div class="step-card"><b>Step 2</b> &nbsp;Transpose — rows become columns, columns become rows</div>
<div class="step-card"><b>Step 3</b> &nbsp;Swap columns A & B</div>
<div class="step-card"><b>Step 4</b> &nbsp;Columns A & B → no fill, black font</div>
<div class="step-card"><b>Step 5</b> &nbsp;Row 1 → bold black | Column B rows 2+ → unbold black</div>
<div class="step-card"><b>Step 6</b> &nbsp;Column B → left align | Columns after Threshold → center align</div>
<div class="step-card"><b>Step 7</b> &nbsp;Auto-size all columns</div>
""", unsafe_allow_html=True)

st.markdown("---")

uploaded_file = st.file_uploader("Upload your Excel file (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        wb_peek = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), read_only=True)
        sheet_names = wb_peek.sheetnames
        wb_peek.close()
        uploaded_file.seek(0)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        st.stop()

    selected_sheet = st.selectbox("Select sheet to process", sheet_names)

    if st.button("▶ Run Transformation", use_container_width=True):
        with st.spinner("Processing..."):
            try:
                file_bytes = uploaded_file.read()
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
                ws = wb[selected_sheet]

                # Step 1: Unmerge all cells and fill every cell in merged range
                merged_ranges = list(ws.merged_cells.ranges)
                for merge in merged_ranges:
                    min_row, min_col = merge.min_row, merge.min_col
                    max_row, max_col = merge.max_row, merge.max_col
                    top_left = ws.cell(row=min_row, column=min_col)
                    val = top_left.value
                    font = copy_obj(top_left.font)
                    fill = copy_obj(top_left.fill)
                    alignment = copy_obj(top_left.alignment)
                    border = copy_obj(top_left.border)
                    number_format = top_left.number_format
                    ws.unmerge_cells(str(merge))
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.value = val
                            cell.font = copy_obj(font)
                            cell.fill = copy_obj(fill)
                            cell.alignment = copy_obj(alignment)
                            cell.border = copy_obj(border)
                            cell.number_format = number_format

                max_row = ws.max_row
                max_col = ws.max_column

                # Step 2: Read all data + styles
                data, styles = [], []
                for r in range(1, max_row + 1):
                    row_vals, row_stls = [], []
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        row_vals.append(cell.value)
                        row_stls.append({
                            "font": copy_obj(cell.font),
                            "fill": copy_obj(cell.fill),
                            "alignment": copy_obj(cell.alignment),
                            "border": copy_obj(cell.border),
                            "number_format": cell.number_format,
                        })
                    data.append(row_vals)
                    styles.append(row_stls)

                # Step 3: Transpose
                transposed_data = list(map(list, zip(*data)))
                transposed_styles = list(map(list, zip(*styles)))

                # Step 4: Write to new Transposed sheet
                if "Transposed" in wb.sheetnames:
                    del wb["Transposed"]
                new_ws = wb.create_sheet("Transposed")

                for r_idx, (row_vals, row_stls) in enumerate(zip(transposed_data, transposed_styles), start=1):
                    for c_idx, (val, stl) in enumerate(zip(row_vals, row_stls), start=1):
                        cell = new_ws.cell(row=r_idx, column=c_idx, value=val)
                        cell.font = copy_obj(stl["font"])
                        cell.fill = copy_obj(stl["fill"])
                        cell.alignment = copy_obj(stl["alignment"])
                        cell.border = copy_obj(stl["border"])
                        cell.number_format = stl["number_format"]

                # Step 5: Swap columns A (1) and B (2) using temp buffer
                col1_buf, col2_buf = [], []
                for r in range(1, new_ws.max_row + 1):
                    c1 = new_ws.cell(row=r, column=1)
                    c2 = new_ws.cell(row=r, column=2)
                    col1_buf.append({
                        "value": c1.value, "font": copy_obj(c1.font), "fill": copy_obj(c1.fill),
                        "alignment": copy_obj(c1.alignment), "border": copy_obj(c1.border),
                        "number_format": c1.number_format,
                    })
                    col2_buf.append({
                        "value": c2.value, "font": copy_obj(c2.font), "fill": copy_obj(c2.fill),
                        "alignment": copy_obj(c2.alignment), "border": copy_obj(c2.border),
                        "number_format": c2.number_format,
                    })

                for r in range(1, new_ws.max_row + 1):
                    c1 = new_ws.cell(row=r, column=1)
                    c2 = new_ws.cell(row=r, column=2)
                    c1.value = col2_buf[r-1]["value"]
                    c1.font = copy_obj(col2_buf[r-1]["font"])
                    c1.fill = copy_obj(col2_buf[r-1]["fill"])
                    c1.alignment = copy_obj(col2_buf[r-1]["alignment"])
                    c1.border = copy_obj(col2_buf[r-1]["border"])
                    c1.number_format = col2_buf[r-1]["number_format"]
                    c2.value = col1_buf[r-1]["value"]
                    c2.font = copy_obj(col1_buf[r-1]["font"])
                    c2.fill = copy_obj(col1_buf[r-1]["fill"])
                    c2.alignment = copy_obj(col1_buf[r-1]["alignment"])
                    c2.border = copy_obj(col1_buf[r-1]["border"])
                    c2.number_format = col1_buf[r-1]["number_format"]

                # Step 6: Find "Threshold" column index in row 1
                threshold_col = None
                for c in range(1, new_ws.max_column + 1):
                    val = new_ws.cell(row=1, column=c).value
                    if val and str(val).strip().lower() == "threshold":
                        threshold_col = c
                        break

                # Step 7: Apply font/fill/alignment rules
                no_fill = PatternFill(fill_type=None)
                for r in range(1, new_ws.max_row + 1):
                    for c in range(1, new_ws.max_column + 1):
                        cell = new_ws.cell(row=r, column=c)

                        if r == 1:
                            # Row 1: bold black for all headers
                            cell.font = Font(
                                name=cell.font.name or "Calibri",
                                size=cell.font.size or 11,
                                bold=True,
                                italic=cell.font.italic,
                                color="FF000000",
                            )
                            if c in (1, 2):
                                cell.fill = no_fill

                        elif c in (1, 2):
                            # Cols A & B rows 2+: no fill, black, unbold
                            cell.fill = no_fill
                            cell.font = Font(
                                name=cell.font.name or "Calibri",
                                size=cell.font.size or 11,
                                bold=False,
                                italic=cell.font.italic,
                                color="FF000000",
                            )

                        # Column B → left align (all rows)
                        if c == 2:
                            cell.alignment = Alignment(
                                horizontal="left",
                                vertical=cell.alignment.vertical,
                                wrap_text=cell.alignment.wrap_text,
                            )

                        # Columns after Threshold → center align (all rows)
                        if threshold_col and c > threshold_col:
                            cell.alignment = Alignment(
                                horizontal="center",
                                vertical=cell.alignment.vertical,
                                wrap_text=cell.alignment.wrap_text,
                            )

                # Step 8: Auto-size all column widths
                for col in new_ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    new_ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                st.markdown('<div class="success-box">✅ Done! Column B left-aligned, data columns center-aligned.</div>', unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)

                original_name = uploaded_file.name.replace(".xlsx", "")
                st.download_button(
                    label="⬇ Download Transposed Excel",
                    data=output,
                    file_name=f"{original_name}_transposed.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error(f"Something went wrong: {e}")
